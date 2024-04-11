# -*- coding: utf-8 -*-
"""
Created on Tue Jun 16 11:57:22 2020

@author: Kumar, Shubham
"""
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
import re
import os
import shutil
import packaging.version
import pandas as pd

class excel_formatting:
    
    header_fill_color = '002E8A'
    header_text_color = 'FFFFFF'
    header_border_color = '000000'
    header_text_size = 11
    
    cell_fill_color = 'FFFFFF'
    cell_text_color = '000000'
    cell_border_color = '000000'
    cell_text_size = 11
    # light blue
    cell_fill_color1 = 'C2D3E8'
    # light purple
    cell_fill_color2 = 'CCC0DA'  
    # light grey
    cell_fill_color3 = 'D9D9D9'
    # light red
    cell_fill_color4 = 'FBA09D'
    # light green
    cell_fill_color5 = 'BFFFBC'
    # light yellow
    cell_fill_color6 = 'FEDE86'
    
    
    title_fill_color = 'A9DA74'
    title_text_color = '1F497D'
    title_border_color = ''
    title_text_size = 14
    
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    def __init__(self):        
        pass
    
    @classmethod
    def reset_header_default(cls):
        version = packaging.version.parse(pd.__version__)
        if version < packaging.version.parse('0.18'):
            pd.pandas.core.format = None
        elif version < packaging.version.parse('0.20'):
            pd.pandas.formats.format = None
        else:
            pd.pandas.io.formats.excel = None
    
    def generate_excel_cell_name(row_number, 
                                 column_number, 
                                 fix_row = False, 
                                 fix_column=False):
        """ This function convert given row and column number to an Excel-style cell name. 
        Parameters
        ----------
        row_number: integer
            Row number (indexing starts from 1)
            
        column_number: integer
            Column number (indexing starts from 1)
        
        fix_row: boolean, default= False
            If true, row is made absolute.
        fix_column: boolean, default= False
            If true, column is made absolute.
        
        """
        result = []
        while column_number:
            column_number, remainder = divmod(column_number-1, 26)
            result.append(excel_formatting.alphabet[remainder])
        result = result[::-1]
        if row_number!=None:
            if fix_row == True:
                if fix_column == True:
                    return '$'+''.join(result) + '$' + str(row_number)
                else:
                    return ''.join(result) + '$' + str(row_number)
            else:
                if fix_column == True:
                    return '$'+''.join(result) + str(row_number)
                else:
                    return ''.join(result) + str(row_number)
        else:
            return ''.join(result)
            

    def generate_excel_range(start_cell,end_cell):
        """ 
        This function convert given start_cell and end_cell to an Excel-style cell range.
    
        Parameters
        ----------
        start_cell: string
            Excel-style cell name of the first cell.
    
        end_cell: integer
            Excel-style cell name of the last cell
        
        Returns
        -------
    
        """
        if start_cell.count('$') == end_cell.count('$'):
            
            start_cell_column, start_cell_row = re.findall('([A-Z]+)\$?([0-9]+)',start_cell)[0]
            end_cell_column, end_cell_row = re.findall('([A-Z]+)\$?([0-9]+)',end_cell)[0]
            
            
            
            start_cell_column = (len(start_cell_column)-1)*26 + excel_formatting.alphabet.index(start_cell_column[-1]) + 1
            end_cell_column = (len(end_cell_column)-1)*26 + excel_formatting.alphabet.index(end_cell_column[-1]) + 1
            if start_cell_column<=end_cell_column:   
                # row number of end cell should be greater tham start cell
                if start_cell_row<=end_cell_row:
                    return start_cell + ':' + end_cell
        else:
            print('Invalid cell range. Please check input values again.')
            return None
    
    @classmethod
    def create_version(cls,
                       base_filename,
                       path=''):
        """This function returns the appropriately versioned folder name so that the existing files do not get overwritten.
        Parameters
        ----------
        base_filename:string
            Base name of the file/folder to be made in the current directory.
            NOTE: Should NOT contain file extension 
        
        path:string
            Absolute path of folder in which to check for existing versions of base file.
        """
        if path == '':
            existing_files = os.listdir()
        else:
            existing_files = os.listdir(path)
        
        if len(existing_files) == 0:
            return base_filename + "_v1"
        else:
            existing_versions = [int(re.findall('\_v([0-9]+)',filename)[-1]) for filename in existing_files if base_filename+"_" in filename]
            if len(existing_versions) == 0:
                return base_filename + "_v1"
            else:
                return base_filename + "_v" + str(max(existing_versions)+1)
 
       
    @classmethod
    def create_folder(cls,
                      folder_name,
                      force_delete_existing = False,
                      path=''):
        
        """This function created the required folder in the specified directory.
        Parameters
        ----------
        folder_name:string
            Name of the folder to be made.
        
        force_delete_existing: boolean, default = False
            If True, deletes existing folder with the same name and makes a new empty folder.
        
        path:string
            Absolute path of folder in which to check for existing versions of base file.
        
        """
        if path == '':
            path = os.getcwd()
            
        try:
            # try to delete existing folder
            shutil.rmtree(os.path.join(path,folder_name))
        except PermissionError:
            # if existing folder is still in use
            print('ERROR:\nFolder '+ folder_name +' already exists and is being used by another process. Please release folder from use.')
            return False
        except FileNotFoundError:
            # if existing folder does not exist
            pass

        try:
            os.mkdir(os.path.join(path,folder_name))
        except PermissionError:
            print('ERROR:\nFolder '+ folder_name +'already exists.')
            #os.mkdir(os.path.join(os.getcwd(),folder_name))
            
    @classmethod
    def formatting_openpyxl(cls, format_name='borders_thin'):
        """This function contains the different formatting styles using openpyxl that can be applied to generate the desired excel output format
        
        Parameters
        ----------
        
        format_name:string, default='borders_thin'
            Name of the preset format types offered by the function to be applied on cells.
        
        Returns
        -------
        style: NamedStyle object
        
        """
        
        if format_name == 'table_header':
            style = NamedStyle(name="table_header")
            style.font = Font(bold=True,color=excel_formatting.header_text_color,size=excel_formatting.header_text_size)
            style.fill = PatternFill("solid", fgColor=excel_formatting.header_fill_color)
            style.border = Border(top = Side(border_style="thin"),
                                  bottom =Side(border_style="thin"),
                                  left = Side(border_style="thin"),
                                  right = Side(border_style="thin"))
            style.alignment = Alignment(horizontal="center", vertical="center")
            return style
        
        elif format_name == 'table_title':
            style = NamedStyle(name="table_title")
            style.font = Font(bold=True,size=excel_formatting.title_text_size,color=excel_formatting.title_text_color)
            style.fill = PatternFill("solid", fgColor=excel_formatting.title_fill_color)
            style.border = Border(top = Side(border_style="thin"),
                                  bottom =Side(border_style="thin"),
                                  left = Side(border_style="thin"),
                                  right = Side(border_style="thin"))
            style.alignment = Alignment(horizontal="left", vertical="center")
            return style
        
        elif format_name == 'table_cells':
            style = NamedStyle(name="table_cells")
            style.border = Border(top = Side(border_style="thin"),
                                  bottom =Side(border_style="thin"),
                                  left = Side(border_style="thin"),
                                  right = Side(border_style="thin"))
            return style
        
        elif format_name == 'cell_highlight1':
            style = NamedStyle(name="cell_highlight1")
            style.font = Font(bold=False,size=excel_formatting.cell_text_size,color=excel_formatting.cell_text_color)
            style.fill = PatternFill("solid", fgColor=excel_formatting.cell_fill_color1)
            style.border = Border(top = Side(border_style="thin"),
                                  bottom =Side(border_style="thin"),
                                  left = Side(border_style="thin"),
                                  right = Side(border_style="thin"))
            return style
            
        else:
            return None
        
        
    @classmethod
    def formatting_xlsxwriter(cls, format_name='cell_odd'):
        """This function contains the different formatting styles using xlsxwriter that can be applied to generate the desired excel output format
        
        Parameters
        ----------
        
        format_name:string, default='cell_odd'
            Name of the preset format types offered by the function to be applied on cells.
        
        Returns
        -------
        style: dictionary
        
        """
        
        if format_name == 'round0':
            style = {'num_format': '0'}
            return style
        
        if format_name == 'round1':
            style = {'num_format': '0.0'}
            return style
        
        if format_name == 'round2':
            style = {'num_format': '0.00'}
            return style
        
        elif format_name == 'round3':
            style = {'num_format': '0.000'}
            return style
        
        elif format_name == 'percsign':
            style = {'num_format': '0.00\%'}
            return style
        
        elif format_name == 'percsignmultipled':
            style = {'num_format': '0.00%'}
            return style
        
        elif format_name == 'border1':
            style = {'bottom':1,'top':5,'left':1,'right':1,
                     'top_color':excel_formatting.title_text_color}
            return style
        
        elif format_name == 'border2':
            style = {'bottom':5,'top':1,'left':1,'right':1,
                     'top_color':excel_formatting.title_text_color}
            return style
        
        elif format_name == 'cell_anomaly_bad':
            style = {'font_size':excel_formatting.cell_text_size,
                     'bg_color': '#'+excel_formatting.cell_fill_color4,
                     'bottom':1,'top':1,'left':1,'right':1}
            return style
        
        elif format_name == 'cell_anomaly_good':
            style = {'font_size':excel_formatting.cell_text_size,
                     'bg_color': '#'+excel_formatting.cell_fill_color5,
                     'bottom':1,'top':1,'left':1,'right':1}
            return style
        
        elif format_name == 'table_title':
            style = {'bold': True, 
                     'font_color':'#'+excel_formatting.title_text_color, 
                     'font_size':excel_formatting.title_text_size,
                     'bg_color': '#'+excel_formatting.title_fill_color,
                     'bottom':1,'top':1,'left':1,'right':1}
            return style
        
        elif format_name == 'table_header':
            style = {'bold': True, 
                     'font_color':'#'+excel_formatting.header_text_color, 
                     'font_size':excel_formatting.header_text_size,
                     'bg_color': '#'+excel_formatting.header_fill_color,
                     'bottom':1,'top':1,'left':1,'right':1}
            return style
        
        elif format_name == 'cell_even':
            style = {'font_size':excel_formatting.cell_text_size,
                     'bg_color': '#'+excel_formatting.cell_fill_color3,
                     'bottom':1,'top':1,'left':1,'right':1}
            return style
        
        elif format_name == 'cell_odd':
            style = {'bottom':1,'top':1,'left':1,'right':1}
            return style

        elif format_name == 'cell_highlight1':
            style = {'bottom':1,'top':1,'left':1,'right':1,
                     'bg_color': '#'+excel_formatting.cell_fill_color1}
            return style
        elif format_name == 'cell_highlight2':
            style = {'bottom':1,'top':1,'left':1,'right':1,
                     'bg_color': '#'+excel_formatting.cell_fill_color2}
            return style
        elif format_name == 'cell_highlight3':
            style = {'bottom':1,'top':1,'left':1,'right':1,
                     'bg_color': '#'+excel_formatting.cell_fill_color3}
            return style
        elif format_name == 'cell_highlight4':
            style = {'bg_color': '#'+excel_formatting.cell_fill_color4}
            return style        
        elif format_name == 'cell_highlight5':
            style = {'bg_color': '#'+excel_formatting.cell_fill_color5}
            return style
        elif format_name == 'cell_highlight6':
            style = {'bg_color': '#'+excel_formatting.cell_fill_color6}
            return style
        elif format_name == 'align_centre':
            style = {'align':'center'}
            return style
        
        else:
            return None
        
        