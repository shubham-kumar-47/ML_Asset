"""
    
    **Project**
    -----------
    Model Tuning
    
    **Description**
    ---------------
    Current version caters to 3 tree based algorithms - Random Forest, GBM and XGBoost.
     
    This module performs ML model tuning in two ways -
    
        1. Searching for optimal parameters using gridsearch
        2. Reducing the features to get the best feature set using 
           recursive feature elimination
    
    This module also has functionality to perform cross validation.
    
    ###**Hyperparameter Tuning** 
    Any ML model needs to be tuned depending on the dataset. This tuning is generally done in terms of its parameters and such parameters are
    called hyperparameters. Tuning of hyperparameters helps in better fitting of the model on the given dataset, which is generally done using 
    gridsearch. *Gridsearch* is the process of searching for optimal parameter value in a grid of possible values for that parameter. The gridsearch 
    can be run for multiple parameter combinations to arrive at an optimal parameter set.
    
    For this, one can use *grid_search_xgboost*, *grid_search_GBM* and *grid_search_random_forest* functions depending on the model.
    
    ###**Recursive Feature Elimination**
    Finding the best set of features is one of the most critical exercises in model development, any important variable should not be left and 
    any unimportant variable should not be used in the model. One always wants to build a model which has minimal variables and high accuracy. This
    is where recursive feature elimination plays a major role. *Recursive Feature Elimination* is a process of removing one variable(least important)
    at a time and then choosing the set of features which gives highest accuracy.
    
    This module contains three functions that perform this task- 
    *recursive_feature_elimination* function optimises any one of the 23 different ML metrics to find the best set features. For example, one can choose to optimise 'F1 score' 
    to see which parameter set provides the best 'F1 score' accuracy.
    
    *recursive_feature_elimination_ks* and *recursive_feature_elimination_lift* functions give the parameter set which optimises risk metrics like test data KS and test data lift
    respectively.
    
    ###**Cross Validation**   
    *Cross Validation* is generally performed to check the stability of the model. K fold cross validation is a process of dividing the data randomly into k equal parts 
    and then training the data on k-1 folds and testing it on the kth fold. Repeating the training and testing k times on k folds, one gets k set of accuracies on k different datasets.
    This shows how the model is performing on different parts of the dataset.
    
    
    Please refer to each class and the corresponding methods to understand their functionality in detail.
    
    **Dependencies**
    ----------------
        Install the following python libraries: 

        numpy==1.19.0
        pandas==1.0.3
        scikit-learn==0.23.1
        xgboost==1.1.0
        openpyxl==3.0.3
        packaging==20.4
        
        
    **Additional Files**
    --------------------
        
        1. validation.py 
        
        2. excel_formatting.py 
    
    Please ensure that the above files are also present in the same folder and packages mentioned in the dependencies section are installed.
    
        
    **Authors**
    -----------
    Created on Wed Jun  10 16:24:13 2020 
    
    @author: Kumar, Shubham

"""

import numpy as np
import os
import re
import pandas as pd
from sklearn.preprocessing import LabelEncoder
import sklearn.metrics as metrics
from sklearn.model_selection import StratifiedKFold
from sklearn.model_selection import KFold
from sklearn.model_selection import ParameterGrid
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingClassifier
from xgboost import XGBClassifier
from src.excel_formatting import excel_formatting as ef
from src.validation import validation as mv

from sklearn.feature_selection import RFECV
from sklearn.svm import SVC
from collections import OrderedDict 
from openpyxl.drawing.image import Image

class model_tuning:
    """
    
    This class contains functions for hyperparameter tuning (RandomForest, GBM, XGBoost), recursive feature selection and cross validation.
    
    """
    
    __save_directory = 'results/Model Tuning'
    if os.path.isdir("results")==False:
        os.mkdir("results")
    
    if os.path.isdir(__save_directory)==False:
        os.mkdir(__save_directory)
        
    # Styles - openpyxl
    __format_table_header = ef.formatting_openpyxl(format_name='table_header')
    __format_table_title = ef.formatting_openpyxl(format_name='table_title')
    __format_table_cells = ef.formatting_openpyxl(format_name='table_cells') 
    __format_cell_highlight1 = ef.formatting_openpyxl(format_name='cell_highlight1')
    
    

        
    def __init__(self):
        pass
    
 
           
    @classmethod
    def __write_cross_validation_results(cls,
                                    filename,
                                    model_parameters_and_metrics,
                                    kfolds):
        """This function writes the cross validation results with the appropriate formatting
        
        Parameters
        ----------
        
        filename: string
            Name of file where results are to be saved.
        
        """ 
        
        filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        filename += '.xlsx'
        try:
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 filename), engine = 'openpyxl',  mode='a')
        except:
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 filename), engine = 'openpyxl')
            
        model_parameters_and_metrics_original = model_parameters_and_metrics.copy()
        
        #print(writer.book.named_styles)
        
        if 'table_header' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_header)
        if 'table_title' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_title)
        if 'table_cells' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_cells)
        if 'cell_highlight1' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_cell_highlight1)
            
        for k in kfolds:
            sheet_name = str(k)+'-fold'
            model_parameters_and_metrics = model_parameters_and_metrics_original[model_parameters_and_metrics_original['kfold']==k]
            # table1
            model_parameters_and_metrics['y_probability_threshold'].to_excel(writer, sheet_name=sheet_name,index=False,startrow=4,startcol=1)
            
            model_metrics = model_parameters_and_metrics.to_dict('records')[0]
            
            metricsTable1 = pd.DataFrame(data=[['Train',model_metrics['train_AUC_ROC'],model_metrics['train_Gini_index'],model_metrics['train_MaxKS']],
                                               ['Test',model_metrics['test_AUC_ROC'],model_metrics['test_Gini_index'],model_metrics['test_MaxKS']]],
                                        columns=['','AUC-ROC','GINI','KS'])
            # table2
            metricsTable1.to_excel(writer, sheet_name=sheet_name,index=False,startrow=8,startcol=1)
            
            metricsTable2 = pd.DataFrame(data=[['Train',model_metrics['train_Accuracy'],model_metrics['train_Precision'],model_metrics['train_Recall'],model_metrics['train_F1_score']],
                                               ['Test',model_metrics['test_Accuracy'],model_metrics['test_Precision'],model_metrics['test_Recall'],model_metrics['test_F1_score']]],
                                        columns=['','Accuracy','Precision','Recall','F1'])
            # table3
            metricsTable2.to_excel(writer, sheet_name=sheet_name,index=False,startrow=13,startcol=1)
            
            metricsTable2 = pd.DataFrame(data=[['Decile1',model_metrics['train_Capture_Decile1'],model_metrics['train_Lift_Decile1'],model_metrics['test_Capture_Decile1'],model_metrics['test_Lift_Decile1'],model_metrics['test_Population%_Decile1']],
                                               ['Decile2',model_metrics['train_Capture_Decile2'],model_metrics['train_Lift_Decile2'],model_metrics['test_Capture_Decile2'],model_metrics['test_Lift_Decile2'],model_metrics['test_Population%_Decile2']],
                                               ['Decile3',model_metrics['train_Capture_Decile3'],model_metrics['train_Lift_Decile3'],model_metrics['test_Capture_Decile3'],model_metrics['test_Lift_Decile3'],model_metrics['test_Population%_Decile3']]],
                                        columns=['','Capture%','Lift','Capture%','Lift','Trigger%'])
            # table4
            metricsTable2.to_excel(writer, sheet_name=sheet_name,index=False,startrow=19,startcol=1)
            
            
            
            # Formatting: Kfold--------------------------------------------------------------
            # -------------------------------------------------------------------------------
            sheet = writer.sheets[sheet_name]
            img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
            img.anchor = ef.generate_excel_cell_name(row_number = 1,
                                                     column_number = 2)
            sheet.add_image(img)
            sheet.cell(row=4, column=2).value = 'Threshold'
            
            sheet.cell(row=8, column=2).value = 'Model Metrics'
            sheet.cell(row=13, column=2).value = 'Model Metrics'  
            sheet.cell(row=18, column=2).value = 'Model Metrics' 
            sheet.cell(row=19, column=3).value = 'Train'  
            sheet.cell(row=19, column=5).value = 'Test'
       
            
            max_column_width = len('y_probability_threshold')
            
            # table headers ------------------------------------
            rows = [5,9,14,19,20] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)] ]
            
                    
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_header'
                    sheet.column_dimensions[column].width = max_column_width 
                    
            # table indices ------------------------------------
            rows = [10,11,15,16,21,22,23] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)] ]
            
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'cell_highlight1'
                    sheet.column_dimensions[column].width = max_column_width 
                    
            # table titles ------------------------------------        
            rows = [4,8,13,18] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)] ]
            
            
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_title'
            
            # table cells ------------------------------------
            rows = [6,10,11,15,16,21,22,23] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,6)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,6)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,7)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,7)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,8)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,8)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(3,8)]]
            
            percentage_columns = ['C','E','F']
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_cells'
                
                if (column in percentage_columns) and (rows[i]>=21):
                    sheet[column+str(rows[i])].number_format = '0.00\%'
                elif i == 0:
                    pass
                else:
                    sheet[column+str(rows[i])].number_format = '0.00'
            
            sheet.merge_cells(start_row=8, start_column=2, end_row=8, end_column=5)
            sheet.merge_cells(start_row=13, start_column=2, end_row=13, end_column=6)
            sheet.merge_cells(start_row=18, start_column=2, end_row=18, end_column=7)
            sheet.merge_cells(start_row=19, start_column=3, end_row=19, end_column=4)
            sheet.merge_cells(start_row=19, start_column=5, end_row=19, end_column=7)
                                
            sheet.sheet_view.showGridLines = False
            # -------------------------------------------------------------------------------
            # -------------------------------------------------------------------------------
            

        writer.save()
        writer.close()
        print("Cross Validation results stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory))
        print("\nFile Name : \n"+ filename)
    
 
    
    @classmethod
    def __find_optimal_cutoff(cls,
                              y_actual,
                              y_predicted_proba):
        
        """
        This function finds the optimal probability cutoff point for a classification model.
        
        Parameters
        ----------
        
        y_actual: array of integers, shape = [n_samples]
            True binary labels.

        y_predicted_proba: array of floats, shape = [n_samples]
            Proba score 
    
        Returns
        -------
        
        threshold: float
            Optimal cutoff value
    
        """
        
        false_positive_rate, true_positive_rate, threshold = metrics.roc_curve(y_actual, y_predicted_proba)
        i = np.arange(len(true_positive_rate)) 
        cutoffs = pd.DataFrame({'younden_index' : pd.Series(true_positive_rate-(1-false_positive_rate), index=i), 'threshold' : pd.Series(threshold, index=i)})
        threshold = cutoffs.iloc[cutoffs.younden_index.abs().argsort()[:1]]['threshold'].tolist()[0]
    
        return threshold
        

    @classmethod
    def __write_grid_search_iterations_results(cls,
                                               filename,
                                               parametersList,
                                               model_metrics,
                                               model_parameters_and_metrics,
                                               feature_importance,
                                               save_folder):
        """This function writes the grid search results with the appropriate formatting
        
        Parameters
        ----------
        
        filename: string
            Name of file without file extension where results are to be saved.
        
        """ 
        
        filename += '.xlsx'
       
        try:
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 save_folder,
                                                 filename), engine = 'openpyxl',  mode='a')
        except:
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 save_folder,
                                                 filename), engine = 'openpyxl')
        
    
        if 'table_header' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_header)
        if 'table_title' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_title)
        if 'table_cells' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_cells)
        if 'cell_highlight1' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_cell_highlight1)
            
        feature_importance.to_excel(writer, sheet_name='FeatureImportance',index=False,startrow=3,startcol=1)
        
        
        # table1
        model_parameters_and_metrics[parametersList].to_excel(writer, sheet_name='ParametersAndMetrics',index=False,startrow=4,startcol=1)
        # table2
        model_parameters_and_metrics['y_probability_threshold'].to_excel(writer, sheet_name='ParametersAndMetrics',index=False,startrow=8,startcol=1)
        
        
        metricsTable1 = pd.DataFrame(data=[['Train',model_metrics['train_AUC_ROC'],model_metrics['train_Gini_index'],
                                            model_metrics['train_MaxKS']],
                                           ['Test',model_metrics['test_AUC_ROC'],model_metrics['test_Gini_index'],
                                            model_metrics['test_MaxKS']]],
                                       columns=['','AUC-ROC','GINI','KS'])
                                     
        # table3
        metricsTable1.to_excel(writer, sheet_name='ParametersAndMetrics',index=False,startrow=12,startcol=1)
        
        metricsTable2 = pd.DataFrame(data=[['Train',model_metrics['train_Accuracy'],model_metrics['train_Precision'],model_metrics['train_Recall'],model_metrics['train_F1_score']],
                                           ['Test',model_metrics['test_Accuracy'],model_metrics['test_Precision'],model_metrics['test_Recall'],model_metrics['test_F1_score']]],
                                    columns=['','Accuracy','Precision','Recall','F1'])
        # table4
        metricsTable2.to_excel(writer, sheet_name='ParametersAndMetrics',index=False,startrow=17,startcol=1)
        
        metricsTable2 = pd.DataFrame(data=[['Decile1',model_metrics['train_Capture_Decile1'],model_metrics['train_Lift_Decile1'],model_metrics['test_Capture_Decile1'],model_metrics['test_Lift_Decile1'],model_metrics['test_Population%_Decile1']],
                                           ['Decile2',model_metrics['train_Capture_Decile2'],model_metrics['train_Lift_Decile2'],model_metrics['test_Capture_Decile2'],model_metrics['test_Lift_Decile2'],model_metrics['test_Population%_Decile2']],
                                           ['Decile3',model_metrics['train_Capture_Decile3'],model_metrics['train_Lift_Decile3'],model_metrics['test_Capture_Decile3'],model_metrics['test_Lift_Decile3'],model_metrics['test_Population%_Decile3']]],
                                    columns=['','Capture%','Lift','Capture%','Lift','Trigger%'])
        # table5
        metricsTable2.to_excel(writer, sheet_name='ParametersAndMetrics',index=False,startrow=23,startcol=1)
        
        
        # Formatting: ParametersAndMetrics--------------------------------------------------------------
        sheet = writer.sheets['ParametersAndMetrics']
        img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
        img.anchor = ef.generate_excel_cell_name(row_number = 1,
                                                 column_number = 2)
        sheet.add_image(img)
        sheet.cell(row=4, column=2).value = 'Model Parameters'
        
        sheet.cell(row=8, column=2).value = 'Threshold'
        sheet.cell(row=12, column=2).value = 'Model Metrics'  
        sheet.cell(row=17, column=2).value = 'Model Metrics' 
        sheet.cell(row=22, column=2).value = 'Model Metrics'  
        sheet.cell(row=23, column=3).value = 'Train'  
        sheet.cell(row=23, column=5).value = 'Test'

        max_column_width = max([len(x) for x in parametersList + ['y_probability_threshold']])
        
        # table headers
        rows = [5,9,13,18,23,24] 
        columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(parametersList)+2)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,7)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)]]
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_header'
                sheet.column_dimensions[column].width = max_column_width 
                
        # table indices
        rows = [14,15,19,20,25,26,27] 
        
        columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)]]
        
                
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'cell_highlight1'
                sheet.column_dimensions[column].width = max_column_width 
                  
        # table titles        
        rows = [4,8,12,17,22] 
        
        columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(parametersList)+2)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,7)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)]]
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_title'
        
        # table cells
        rows = [6,10,14,15,19,20,25,26,27] 
        columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(parametersList)+2)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,6)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,7)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,7)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)],
                    [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,8)]]

       
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_cells'
                
                if (column in ['C','E','F']) and (rows[i]>=25):
                    sheet[column+str(rows[i])].number_format = '0.00\%'
                elif i == 0:
                    pass
                else:
                    sheet[column+str(rows[i])].number_format = '0.00'
        
        sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=len(parametersList)+1)
        sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
        sheet.merge_cells(start_row=17, start_column=2, end_row=17, end_column=6)
        sheet.merge_cells(start_row=22, start_column=2, end_row=22, end_column=7)
        sheet.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
        sheet.merge_cells(start_row=23, start_column=5, end_row=23, end_column=7)
                
        
        sheet.sheet_view.showGridLines = False
        # -------------------------------------------------------------------------------------------
        # Formatting: FeatureImportance -------------------------------------------------------------
        # -------------------------------------------------------------------------------------------
        
        sheet = writer.sheets['FeatureImportance']
        img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
        img.anchor = 'B1'
        sheet.add_image(img)
        
        max_column_width = max([len(x) + 2 for x in feature_importance['Variable']])
        
        # table headers ----------------------------------------------------------------------------- 
        rows = [2] 
        columns = [[ef.generate_excel_cell_name(None,column_number) for column_number in range(2,3)]]

        
        sheet['B4'].style = 'table_header'
        sheet.column_dimensions['B'].width = max_column_width 
        sheet['C4'].style = 'table_header'
        sheet.column_dimensions['C'].width = len('VariableImportance') + 3
                  
        # table cells -------------------------------------------------------------------------------
        rows = range(5,len(feature_importance)+5)
        columns = [[ef.generate_excel_cell_name(None,column_number) for column_number in range(2,4)]]*len(feature_importance)
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_cells'

                if column == 'C':
                    sheet[column+str(rows[i])].style = 'table_cells'
                    sheet[column+str(rows[i])].number_format = '0.00%'
                else:
                    sheet[column+str(rows[i])].style = 'cell_highlight1'
                
        
        sheet.sheet_view.showGridLines = False
        # ---------------------------------------------------------------------
        
        writer.save()
        writer.close()
        
    @classmethod
    def __model_metrics(cls,
                        model,
                        x_train,x_test,
                        y_train,y_test,
                        iteration_number,
                        test_calculation_method,
                        include_plots,
                        save_folder,
                        score_cutoff=None,
                        save_charts=True):
        
        """This function calculates the model performance metrics such as AUC_ROC, Accuracy, Precision, Recall, F1_score, Gini_index,
            Capture_Decile1, Capture_Decile2, Capture_Decile3, MaxKS, MaxLift, Lift_Decile1.
        
        Parameters
        ----------
        
        model: estimator object
            Model object for the ml algorithm used.
            
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
    		Training data, where n_samples is the number of samples and n_features is the number of features.
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
    		Test data, where n_samples is the number of samples and n_features is the number of features.
        
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
                        
        iteration_number: integer
                          grid search iteration number
            
        test_calculation_method: string, default="ScoreCutOff"
            Whether to divide test set into equal size deciles or divide into deciles using score cuts.
            Accepted values: {"ScoreCutOff", "EqualDeciles"}.
        
        include_plots: boolean,
            True, to save KS Lift plot for every iteration.
        
        save_folder: string
                     Folder name where the output files must be saved
                     By default will save files in the same location as the code.
        
        score_cutoff: float, default=None
            Optimal cutoff score for y probabilities.
            If None, will calculate optimal score cut off
                       
        save_charts: boolean
                     default = True
                     True, to save KS charts for every iteration.
        
        Returns
        -------
        model_metrics: OrderedDict
            Contains following metrics for both test train - AUC_ROC, Accuracy, Precision, Recall, F1_score, Gini_index,
            Capture_Decile1, Capture_Decile2, Capture_Decile3, MaxKS, MaxLift, Lift_Decile1
         
        """
        
        # Predicted probabilites
        y_train_predicted_proba = model.predict_proba(x_train)[:,1]
        y_test_predicted_proba = model.predict_proba(x_test)[:,1]

        # Optimal cut off for y
        if score_cutoff == None:
            threshold = cls.__find_optimal_cutoff(y_test, y_test_predicted_proba)
        else:
            threshold = score_cutoff

        y_train_predicted = np.where(y_train_predicted_proba >= threshold, 1, 0)
        y_test_predicted = np.where(y_test_predicted_proba >= threshold, 1, 0)

        
        y_train_probability = pd.DataFrame({'Target': y_train,"Probability":y_train_predicted_proba})

        
        y_test_probability = pd.DataFrame({'Target': y_test,"Probability":y_test_predicted_proba})


        # KS for Train and Test
        (KS_train,KS_test) = mv.calculate_ks(y_train_probability,
                                             y_test_probability,
                                             iteration_number=iteration_number,
                                             test_calculation_method=test_calculation_method,
                                             include_plots=include_plots,
                                             save_folder=save_folder,
                                             save_charts=save_charts)
        
        empty_df = pd.DataFrame(index= range(len(KS_test)+1,11),columns=list(KS_test.columns))
        KS_test = pd.concat([KS_test,empty_df])
        KS_train["dBadRate"]=KS_train["BadRate"].diff(-1)
        KS_test["dBadRate"]=KS_test["BadRate"].diff(-1)
 
        #concordance_disconcordance_matrices_train=concordance_disconcordance(x_train,y_train,model)
        #concordance_disconcordance_matrices_test=concordance_disconcordance(x_test,y_test,model)
        aa=list(KS_train['dBadRate'])
        rt=[aa.index(va)+2 for va in aa if va<0]
        if rt==[]:
            train_break='no-break'
        else:
            train_break=rt[0]
            
        aaa=list(KS_test['dBadRate'])
        rtt=[aaa.index(va)+2 for va in aaa if va<0]
        if rtt==[]:
            test_break='no-break'
        else:
            test_break=rtt[0]

        model_metrics = OrderedDict({'y_probability_threshold':threshold,
                                     'train_AUC_ROC': metrics.roc_auc_score(y_train,y_train_predicted),
                                     'test_AUC_ROC': metrics.roc_auc_score(y_test,y_test_predicted),
                                     'train_Accuracy':metrics.accuracy_score(y_train,y_train_predicted),
                                     'test_Accuracy':metrics.accuracy_score(y_test,y_test_predicted),
                                     'train_Precision':metrics.precision_score(y_train,y_train_predicted),
                                     'test_Precision':metrics.precision_score(y_test,y_test_predicted),
                                     'train_Recall':metrics.recall_score(y_train,y_train_predicted),
                                     'test_Recall':metrics.recall_score(y_test,y_test_predicted),
                                     'train_F1_score':metrics.f1_score(y_train,y_train_predicted),
                                     'test_F1_score':metrics.f1_score(y_test,y_test_predicted),
                                     'train_Gini_index':(2*metrics.roc_auc_score(y_train,y_train_predicted) - 1),
                                     'test_Gini_index':(2*metrics.roc_auc_score(y_test,y_test_predicted) - 1),
                                     'train_break':train_break,
                                     'test_break':test_break,
                                     
                                     
                                    #'train_percent_concordant':concordance_disconcordance_matrices_train[0],
                                    #'test_percent_concordant':concordance_disconcordance_matrices_test[0],
                                    #'train_percent_discordant':concordance_disconcordance_matrices_train[1],
                                    #'test_percent_discordant':concordance_disconcordance_matrices_test[1],
                                    #'train_percent_tied':concordance_disconcordance_matrices_train[2],
                                    #'test_percent_tied':concordance_disconcordance_matrices_test[2],
                
                                    'train_Capture_Decile1':KS_train.loc[1,"Cum_Event%"],
                                    'test_Capture_Decile1':KS_test.loc[1,"Cum_Event%"],
                                    'train_Capture_Decile2':KS_train.loc[2,"Cum_Event%"],
                                    'test_Capture_Decile2':KS_test.loc[2,"Cum_Event%"],
                                    'train_Capture_Decile3':KS_train.loc[3,"Cum_Event%"],
                                    'test_Capture_Decile3':KS_test.loc[3,"Cum_Event%"],
                                    'train_MaxKS':KS_train['KS'].max(),
                                    'test_MaxKS':KS_test['KS'].max(),
                                    'train_MaxLift':KS_train['Lift'].max(),
                                    'test_MaxLift':KS_test['Lift'].max(),
                                    'train_Lift_Decile1':KS_train.loc[1,"Lift"],
                                    'test_Lift_Decile1':KS_test.loc[1,"Lift"],
                                    'train_Lift_Decile2':KS_train.loc[2,"Lift"],
                                    'test_Lift_Decile2':KS_test.loc[2,"Lift"],
                                    'train_Lift_Decile3':KS_train.loc[3,"Lift"],
                                    'test_Lift_Decile3':KS_test.loc[3,"Lift"],
                                    
                                    'train_Population%_Decile1':KS_train.loc[1,"Population%"],
                                    'test_Population%_Decile1':KS_test.loc[1,"Population%"],
                                    'train_Population%_Decile2':KS_train.loc[2,"Population%"],
                                    'test_Population%_Decile2':KS_test.loc[2,"Population%"],
                                    'train_Population%_Decile3':KS_train.loc[3,"Population%"],
                                    'test_Population%_Decile3':KS_test.loc[3,"Population%"]})
                

        return model_metrics

        
    @classmethod
    def cross_validation(cls,
                         x, y, 
                         model_object,
                         score_cutoff,
                         kfolds = [5,10], 
                         random_state = 0,
                         method_type='StratifiedKFold',
                         test_calculation_method="ScoreCutOff"):
        
        """
        This function performs k-fold cross validation.
        
        Parameters
        ----------
        
        x: array-like, sparse matrix of shape (n_samples, n_features)
    		Training data, where n_samples is the number of samples and n_features is the number of features.
        
        y: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable for supervised learning problems.
        
        model_object: estimator object
            Model object for the ml algorithm used.
            
        score_cutoff: float
            Optimal cutoff score for y probabilities.
        
        kfolds: list of integers, default = [5]
            Number of folds. Must be at least 2.
            
        random_state: integer, default=0
            Random number seed.
        
        method_type: string, default = 'StratifiedKFold'
            Specifies which cross-validation function to use.
            Accepted values: {"StratifiedKFold", "KFold"}.
        
        test_calculation_method: string, default="ScoreCutOff"
            Whether to divide test set into equal size deciles or divide into deciles using score cuts.
            Accepted values: {"ScoreCutOff", "EqualDeciles"}.
        
        Returns
        -------
        
        cross_validation_scores: dataframe
            Contains average metric scores for every k fold iteration
        
        Additional files generated
        --------------------------
        
        modelname_CrossValidationScores.csv: csv file containing cross_validation_scores
        
        Example
        -------
        >>> from model_tuning import model_tuning
        >>> modelObj = RandomForestClassifier()
        >>> cross_validation_scores = model_tuning.cross_validation(x, y, modelObj, model_name='RandomForest')
		 
		"""

        metric_value_list = []
        
        if (np.isfinite(x).values.all()==False) or (np.isfinite(y).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
        try:
            print('Running iterations.......................\n')
            
            model_name = str(model_object)
            model_name = re.findall('([a-zA-Z]+)Classifier',model_name)[0]
            
            filename = model_name+"_CrossValidationScores"
            # versioning of filename is handled internally in __write_cross_validation_results(), which is called below
            
            for k in kfolds:
                
                if method_type == 'StratifiedKFold':
                    cv = StratifiedKFold(n_splits=k, random_state=random_state, shuffle=True)
                if method_type == 'KFold':
                    cv = KFold(n_splits=k, random_state=random_state, shuffle=True)
                
                print('Performing',k,'Fold...')
                
                iteration_number = 0
                for train_index, test_index in cv.split(x,y):
                    iteration_number +=1
    
                    x_train, x_test = x.iloc[train_index], x.iloc[test_index]
                    y_train, y_test = y.iloc[train_index], y.iloc[test_index]
                    model_object.fit(x_train, y_train)
    
                    
                    model_metrics = OrderedDict({'kfold':k,
                                    'setNumber':iteration_number})
                                    
                    model_metrics.update(cls.__model_metrics(model=model_object,
                                                              x_train=x_train,x_test=x_test,
                                                              y_train=y_train,y_test=y_test,
                                                              iteration_number=iteration_number,
                                                              test_calculation_method=test_calculation_method,
                                                              save_charts=False,
                                                              score_cutoff = score_cutoff,
                                                              include_plots=False,
                                                              save_folder=''))
                    
                    
                    metric_value_list.append(model_metrics)
    
            metric_values = pd.DataFrame(metric_value_list)            
            cross_validation_scores = metric_values.groupby(["kfold"], as_index=False).mean()
            cross_validation_scores.drop(columns = ["setNumber"], inplace=True)
            cls.__write_cross_validation_results(filename=filename,
                                                model_parameters_and_metrics = cross_validation_scores,
                                                kfolds = kfolds)
            
            return cross_validation_scores
        except TypeError as te:
            print('ERROR:Incorrect datatype of one or more parameters entered. Please check the documentation for detailed understanding.')
            print('ERROR:',str(te))
        except ValueError as ve:
            print('ERROR:',str(ve))
            
    
    @classmethod
    def recursive_feature_elimination(cls,
                                      data,
                                      target_name,
                                      method,
                                      score_type):
        """
        This function performs recursive feature elimination using the selected ML method(SVC or RF).
        It chooses the features based on metric choosen by the user.
                
        Parameters
        ----------
        
        data: dataset
            Data containing both x and y variables.
        
        target_name: string
            Name of the column containing target (y) variable.
        
        method: string
            Name of ML model which is used to perform rfe.
            Accepted values: {"SVC", "RandomForest"}.
        
        score_type: string
            Performance metrics to be used for feature reduction
            Accepted values: {'accuracy','balanced_accuracy','precision','recall'
                              'average_precision','neg_brier_score','neg_log_loss'
                              'f1','f1_micro','f1_macro','f1_weighted','f1_samples',
                              'jaccard','roc_auc','roc_auc_ovr','roc_auc_ovo',
                              'roc_auc_ovr_weighted','roc_auc_ovo_weighted'}
        Returns
        -------
        
        feature_ranking: dataframe
            Ranking of each feature with its corresponding importance
        
        Example
        -------
        
        >>> from model_tuning import model_tuning
        >>> feature_ranking = model_tuning.recursive_feature_elimination(data, target_name='Survived', method='RandomForest', score_type="accuracy")
		
        """
        
        data_copy = data.copy()
        if np.isfinite(data).values.all()==False:
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
	             
        x_column_names = [x for x in data.columns if x not in target_name]
        if method=='SVC':
            model_object = SVC(kernel="linear")
        elif method=='RandomForest':
            model_object = RandomForestClassifier(random_state=0, n_estimators=int(np.sqrt(data.shape[1]))) 
        else:
            raise Exception('ERROR:Incorrect value of method entered. Please check the documentation for accepted values.')
        selector = RFECV(model_object, step=1, cv=5, scoring=score_type)
        selector = selector.fit(data[x_column_names], data[target_name])
        
        feature_ranking=pd.DataFrame({'Variable': x_column_names, 'Ranking': selector.ranking_})
        feature_ranking.sort_values(by=['Ranking','Variable'],ascending=True,inplace=True)
        
        model_name = str(model_object)
        model_name = re.findall('([a-zA-Z]+)',model_name)[0]
        filename = model_name+"_RecursiveFeatureElimination_"+score_type
        filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        filename += '.xlsx'
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        
        writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                             model_tuning.__save_directory,
                                             filename), engine = 'openpyxl')
        
        if 'table_header' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_header)
        if 'table_title' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_title)
        if 'table_cells' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_cells)
        if 'cell_highlight1' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
        feature_ranking.to_excel(writer, sheet_name='Ranking',index=False,startrow=3,startcol=1)
        
        # Formatting: Ranking--------------------------------------------------------------
        sheet = writer.sheets['Ranking']
        img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
        img.anchor = 'B1'
        sheet.add_image(img)
        
        max_column_width = max([len(x) + 2 for x in feature_ranking['Variable']])
        
        # table headers        
        rows = [2] 
        columns = [list(alphabet[:2])]
        
        sheet['B4'].style = 'table_header'
        sheet.column_dimensions['B'].width = max_column_width 
        sheet['C4'].style = 'table_header'
        sheet.column_dimensions['C'].width = len('Ranking') + 3
                  
        # table cells
        rows = range(5,len(feature_ranking)+5)
        columns = [list(alphabet[1:3])]*len(feature_ranking)
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_cells'

                if column == 'C':
                    sheet[column+str(rows[i])].style = 'table_cells'
                else:
                    sheet[column+str(rows[i])].style = 'cell_highlight1'
                
        
        sheet.sheet_view.showGridLines = False
        # ---------------------------------------------------------------------
        
        
        writer.save()
        writer.close()
        
        print("Recursive Feature Elimination results stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory))
        print("\nFile Name : \n"+ filename)
        
        feature_ranking.reset_index(drop=True,inplace=True)
        data = data_copy
        return(feature_ranking)
        
    @classmethod
    def recursive_feature_elimination_ks(cls,
                                         x_train,
                                         y_train,
                                         x_test,
                                         y_test,
                                         model_object):
        """
        This function performs recursive feature elimination based on test data KS.
        
        Parameters
        ----------
        
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
    		Training data, where n_samples is the number of samples and n_features is the number of features.
        
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
    		Test data, where n_samples is the number of samples and n_features is the number of features.            .
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
        
        model_object: estimator object
            Model object for the ml algorithm used.
        
        Returns
        -------
        
        result: dataframe
            Dataframe containing max KS, 1st decile capture, 2nd decile capture, 3rd decile capture for train and test dataset and removed column name,
            which contains the name of the additionally removed variable for that iteration. 
            (Note: All the variables listed i-1 row, will have also been removed for the ith row iteration)
        
        selected_columns: list of strings
            List of optimal feature names
        
        model_object: estimator object
            Model object for the ml algorithm trained on data containing only the optimal features
        
        
        Example
        -------
        >>> from model_tuning import model_tuning
        >>> model = XGBClassifier()
        >>> result,selected_columns,model_object = model_tuning.recursive_feature_elimination_ks(x_train,y_train,x_test,y_test,model)
        
        Usage
        -------   
        It is recommended to perform this step once the user has a first cut model with 20 or less variables.
        This method will help the user to find the optimal set of features from the first cut model. 
        
        """
        
        if (np.isfinite(x_train).values.all()==False) or (np.isfinite(y_train).values.all()==False) or (np.isfinite(x_test).values.all()==False) or (np.isfinite(y_test).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
        
        # keeping the original dataset so that model can be fit on optimal set of features at the end
        x_train_original = x_train.copy()
        
        test_ks =[]
        train_ks =[]
        train_capture_decile1=[]
        train_capture_decile2=[]
        train_capture_decile3=[]
        test_capture_decile1=[]
        test_capture_decile2=[]
        test_capture_decile3=[]

        
        removed_column = ['']
        resultDict = {}
        max_ks = 0
        highlight_index = 0
        for x_column in x_train.columns:
            model_object.fit(x_train,y_train)
            
            predictors=x_train.columns
            feature_importance = pd.Series(data=model_object.feature_importances_, index=predictors)
            
            # Predicted probabilites
            y_train_predicted_proba = model_object.predict_proba(x_train)[:,1]
            y_test_predicted_proba = model_object.predict_proba(x_test)[:,1]
            
            y_train_probability = pd.DataFrame({'Target': y_train,"Probability":y_train_predicted_proba})
            y_test_probability = pd.DataFrame({'Target': y_test,"Probability":y_test_predicted_proba})
            
            (KS_train,KS_test) = mv.calculate_ks(y_train_probability,y_test_probability,
                                                 iteration_number=1,
                                                 include_plots=False,
                                                 save_folder='',
                                                 save_charts=False)
            
            
            train_ks.append(KS_train['KS'].max())
            test_ks.append(KS_test['KS'].max())
            train_capture_decile1.append(KS_train.loc[1,"Cum_Event%"])
            train_capture_decile2.append(KS_train.loc[2,"Cum_Event%"])
            train_capture_decile3.append(KS_train.loc[3,"Cum_Event%"])
            test_capture_decile1.append(KS_test.loc[1,"Cum_Event%"])
            test_capture_decile2.append(KS_test.loc[2,"Cum_Event%"])
            test_capture_decile3.append(KS_test.loc[3,"Cum_Event%"])
                
            if len(x_train.columns) > 5:
                # for the next iteration, select worst variable to drop
                drop_column = feature_importance.idxmin()
                #print("columns_to_drop",feature_importance.idxmin())
            else:
                break

            ks = KS_test['KS'].max()
            if ks >= max_ks:
                max_ks = ks
                selected_columns = x_train.columns

            removed_column.append(drop_column)
            x_train = x_train.drop([drop_column],axis=1)
            x_test = x_test.drop([drop_column],axis=1)
            
        resultDict['Removed Variable'] = removed_column
        resultDict['Test KS'] = test_ks
        resultDict['Train KS'] = train_ks
        resultDict['Train Decile1 Capture'] = train_capture_decile1
        resultDict['Train Decile2 Capture'] = train_capture_decile2
        resultDict['Train Decile3 Capture'] = train_capture_decile3
        resultDict['Test Decile1 Capture'] = test_capture_decile1
        resultDict['Test Decile2 Capture'] = test_capture_decile2
        resultDict['Test Decile3 Capture'] = test_capture_decile3
        
        result = pd.DataFrame(resultDict)        
        highlight_index = max(result[result['Test KS']==max_ks].index)
        result[['Test KS','Train KS','Train Decile1 Capture','Train Decile2 Capture','Train Decile3 Capture','Test Decile1 Capture','Test Decile2 Capture','Test Decile3 Capture']] = result[['Test KS','Train KS','Train Decile1 Capture','Train Decile2 Capture','Train Decile3 Capture','Test Decile1 Capture','Test Decile2 Capture','Test Decile3 Capture']].round(decimals=3)
        
        model_object.fit(x_train_original[selected_columns],y_train)
        
        model_name = str(model_object)
        model_name = re.findall('([a-zA-Z]+)',model_name)[0]
        filename = model_name+"_RecursiveFeatureElimination_KS"
        filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        filename += '.xlsx'
        
        
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        
        writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                             model_tuning.__save_directory,
                                             filename), engine = 'openpyxl')
        
        if 'table_header' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_header)
        if 'table_title' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_title)
        if 'table_cells' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_cells)
        if 'cell_highlight1' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
        result.to_excel(writer, sheet_name='RecursiveFeatureElimination',index=False,startrow=3,startcol=1)
        
        # Formatting: FeatureImportance--------------------------------------------------------------
        sheet = writer.sheets['RecursiveFeatureElimination']
        img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
        img.anchor = 'B1'
        sheet.add_image(img)
        
        max_column_width = max([len(x) + 2 for x in result['Removed Variable']])
        
        # table headers        
        rows = [4] 
        columns = [list(alphabet[1:len(result.columns)+1])]
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_header'
                if column == 'B':
                    sheet.column_dimensions[column].width = max_column_width 
                else:
                    sheet.column_dimensions[column].width = len(sheet[column+str('4')].value) + 2
                  
        # table cells
        rows = range(5,len(result)+5)
        columns = [list(alphabet[1:len(result.columns)+1])]*len(result)
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_cells'
                if i == highlight_index:
                    sheet[column+str(rows[i])].style = 'cell_highlight1'
                    
                if column in ['C','D']:
                    sheet[column+str(rows[i])].number_format = '0.000'
                else:
                    sheet[column+str(rows[i])].number_format = '0.00\%'
        
        sheet.sheet_view.showGridLines = False
        # ---------------------------------------------------------------------
        
        
        writer.save()
        writer.close()
        
        print("Recursive Feature Elimination results stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory))
        print("\nFile Name : \n"+ filename)     
        
        return result, selected_columns, model_object
    
    
    
    @classmethod
    def recursive_feature_elimination_lift(cls,
                                         x_train,
                                         y_train,
                                         x_test,
                                         y_test,
                                         model_object):
        """
        This function performs recursive feature elimination based on max test data first decile lift.
        
        Parameters
        ----------
        
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
    		Training data, where n_samples is the number of samples and n_features is the number of features.
        
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
    		Test data, where n_samples is the number of samples and n_features is the number of features.            .
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
        
        model_object: estimator object
            Model object for the ml algorithm used.
        
        Returns
        -------
        
        result: dataframe
            Dataframe containing 1st decile lift for train and test dataset and removed column name,
            which contains the name of the additionally removed variable for that iteration. 
            (Note: All the variables listed i-1 row, will have also been removed for the i row iteration)
        
        selected_columns: list of strings
            List of optimal feature names
        
        model_object: estimator object
            Model object for the ml algorithm trained on data containing only the optimal features
        
        Example
        -------
        >>> from model_tuning import model_tuning
        >>> model = XGBClassifier()
        >>> result, selected_columns, model_object = model_tuning.recursive_feature_elimination_ks(x_train,y_train,x_test,y_test,model)
        
        Usage
        -------   
        It is recommended to perform this step once the user has a first cut model with 20 or less variables.
        This method will help the user to find the optimal set of features from the first cut model.
        
        """
        
        if (np.isfinite(x_train).values.all()==False) or (np.isfinite(y_train).values.all()==False) or (np.isfinite(x_test).values.all()==False) or (np.isfinite(y_test).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
        
        # keeping the original dataset so that model can be fit on optimal set of features at the end
        x_train_original = x_train.copy()
        
    
        train_lift_decile1=[]
        test_lift_decile1=[]

        
        removed_column = ['']
        resultDict = {}
        max_lift = 0
        
        for x_column in x_train.columns:
            model_object.fit(x_train,y_train)
            
            predictors=x_train.columns
            feature_importance = pd.Series(data=model_object.feature_importances_, index=predictors)
            
            # Predicted probabilites
            y_train_predicted_proba = model_object.predict_proba(x_train)[:,1]
            y_test_predicted_proba = model_object.predict_proba(x_test)[:,1]
            
            y_train_probability = pd.DataFrame({'Target': y_train,"Probability":y_train_predicted_proba})
            y_test_probability = pd.DataFrame({'Target': y_test,"Probability":y_test_predicted_proba})
            
            (KS_train,KS_test) = mv.calculate_ks(y_train_probability,y_test_probability,
                                                 iteration_number=1,
                                                 include_plots=False,
                                                 save_folder='',
                                                 save_charts=False)
            
            
            train_lift_decile1.append(KS_train.loc[1,"Lift"])
            test_lift_decile1.append(KS_test.loc[1,"Lift"])
                
            if len(x_train.columns) > 5:
                # for the next iteration, select worst variable to drop
                drop_column = feature_importance.idxmin()
                #print("columns_to_drop",feature_importance.idxmin())
            else:
                break

            lift = KS_test.loc[1,"Lift"]
            if lift >= max_lift:
                max_lift = lift
                selected_columns = x_train.columns

            removed_column.append(drop_column)
            x_train = x_train.drop([drop_column],axis=1)
            x_test = x_test.drop([drop_column],axis=1)
            
        resultDict['Removed Variable'] = removed_column
        resultDict['Train Decile1 Lift'] = train_lift_decile1
        resultDict['Test Decile1 Lift'] = test_lift_decile1
        
        result = pd.DataFrame(resultDict)
        highlight_index = max(result[result['Test Decile1 Lift']==max_lift].index)
        result[['Train Decile1 Lift','Test Decile1 Lift']] = result[['Train Decile1 Lift','Test Decile1 Lift']].round(decimals=3)
        
        model_object.fit(x_train_original[selected_columns],y_train)
        
        
        
        model_name = str(model_object)
        model_name = re.findall('([a-zA-Z]+)',model_name)[0]
        filename = model_name+"_RecursiveFeatureElimination_Lift"
        filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        filename += '.xlsx'
        
        
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        
        writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                             model_tuning.__save_directory,
                                             filename + ".xlsx"), engine = 'openpyxl')
        if 'table_header' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_header)
        if 'table_title' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_title)
        if 'table_cells' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_table_cells)
        if 'cell_highlight1' not in writer.book.named_styles:
            writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
        result.to_excel(writer, sheet_name='RecursiveFeatureElimination',index=False,startrow=3,startcol=1)
        
        # Formatting: RecursiveFeatureElimination--------------------------------------------------------------
        
        sheet = writer.sheets['RecursiveFeatureElimination']
        img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
        img.anchor = 'B1'
        sheet.add_image(img)
        max_column_width = max([len(x) + 2 for x in result['Removed Variable']])
        
        # table headers        
        rows = [4] 
        columns = [list(alphabet[1:len(result.columns)+1])]
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_header'
                if column == 'B':
                    sheet.column_dimensions[column].width = max_column_width 
                else:
                    sheet.column_dimensions[column].width = len(sheet[column+str('4')].value) + 2
                  
        # table cells
        rows = range(5,len(result)+5)
        columns = [list(alphabet[1:len(result.columns)+1])]*len(result)
        
        for i in range(0,len(rows)):
            for column in columns[i]:
                sheet[column+str(rows[i])].style = 'table_cells'
                if i == highlight_index:
                    sheet[column+str(rows[i])].style = 'cell_highlight1'
                    
                if column in ['C','D']:
                    sheet[column+str(rows[i])].number_format = '0.000'
                else:
                    sheet[column+str(rows[i])].number_format = '0.00\%'


                
                    
        
        writer.save()
        writer.close()
        
        print("Recursive Feature Elimination results stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory))
        print("\nFile Name : \n"+ filename)     
        
        return result, selected_columns, model_object
    
    
    @classmethod
    def grid_search_xgboost(cls,
                            x_train,x_test,
                            y_train,y_test,
                            subsample= [0.6],learning_rate=[0.01],
                            max_depth=[3],n_estimators=[600],
                            min_child_weight= [1],gamma= [0.5],
							colsample_bytree=[1],
                            random_state = 0,
                            test_calculation_method = "ScoreCutOff",
                            save_charts = True,
                            include_plots = True):
    

        """This function performs grid search for hyperparameter tuning for XGBClassifier.
        
        Please note that by default this function creates a csv for every iteration and hence may be time consuming.
        Set the save_charts flag to False to disable the save feature. 
        
        Parameters
        ----------
        
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
    		Training data, where n_samples is the number of samples and n_features is the number of features.
            .
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
    		Test data, where n_samples is the number of samples and n_features is the number of features.
            .
        
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
            .
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
            .
        
        max_depth: list of integers, default=[3]
            Proposed range=[3,4,5].
            Maximum tree depth for base learners.
        
        learning_rate: list of floats, default=[0.1]
            Proposed range=[0.02,0.04,0.06,0.08,0.1].
            Boosting learning rate (xgb's 'eta').
        
        n_estimators: list of integers, default=[100]
            Proposed range=[600,800,1000,1200].
            Number of trees to fit.
        
        gamma: list of floats, default=[0]
            Proposed range=[0.5,1,1,5,2].
            Minimum loss reduction required to make a further partition on a leaf node of the tree.
        
        min_child_weight: list of floats, default=[1]
            Proposed range=[0.8,1].
            Minimum sum of instance weight(hessian) needed in a child.
        
        subsample: list of floats, default=[1]
            Proposed range=[0.6, 0.8, 1.0].
            Subsample ratio of the training instance.
        
        random_state: integer, default=0
            Random number seed.
        
        test_calculation_method: string, default="ScoreCutOff"
            Whether to divide test set into equal size deciles or divide into deciles using score cuts.
            Accepted values: {"ScoreCutOff", "EqualDeciles"}.
            
        save_charts: boolean
            default = True
            True, to save KS charts for every iteration.
        
        include_plots: boolean, default = True
            True, to save KS Lift plot for every iteration.
            
        Returns
        -------
        
        grid_search_results: dataframe
            Contains parameter values and metric scores for every iteration.
            
        model_object_list: list
            Contains the model objects generated for each iteration
        
        Additional files generated
        --------------------------
        Iteration folder: KS charts for every iteration.
        
        XGB_GridSearchResults.csv: csv file containing grid_search_results.

        
        Usage
        -----
        This functions results one summary csv 'GradientBoosting_GridSearchResults' which has all the parameter sets and the corresponding accuracies. One can select the best parameter set
        based on the accuracy results. Once the optimal parameter set is chosen, one can look at all the results and plots related to the same iteration 
        in the 'Iteration folder' which stores results for each iteration in a seperate file. So let's say iteration 10 is the optimal one, iteration 10 file can be 
        used to consume the detailed formatted results related to that iteration. Model object with optimal parameters can also be selected from the 'model_object_list'
        for further analysis. So if 10th iteration is the optimal one, one can select 10th element of this list to get the corresponding model object.
       
        
        Example
        -------
        
        >>> from model_tuning import model_tuning
        >>> grid_search_results, model_object_list = model_tuning.grid_search_xgboost(x_train, x_test, y_train, y_test, max_depth=[3,4,5])
        
            
        """
        
        if (np.isfinite(x_train).values.all()==False) or (np.isfinite(y_train).values.all()==False) or (np.isfinite(x_test).values.all()==False) or (np.isfinite(y_test).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
            
        algorithm_name = 'XGB'
        save_folder = algorithm_name+'_'+'Iterations'
        save_folder = ef.create_version(base_filename=save_folder, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        status = ef.create_folder(save_folder,path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        if status==False:
            return
    


        try:
            label_encoder_object = LabelEncoder()
            y_train = label_encoder_object.fit_transform(y_train)
            y_test = label_encoder_object.fit_transform(y_test)
            
            
            
            # A list of (X, y) tuple pairs to use as validation sets,
            # for which metrics will be computed which will be used to decide the early_stopping_rounds
            eval_set = [(x_train,y_train)]
            scale_pos_weight=(len(y_train)-y_train.sum())/y_train.sum()
            
            
            iteration_number = 0
            grid_search_list = []
            model_object_list = []
    
    
            print("Running iterations..........\n")
            
            param_grid = {'subsample':subsample,
                          'learning_rate':learning_rate,
                          'max_depth':max_depth,
                          'n_estimators':n_estimators,
                          'min_child_weight':min_child_weight,
                          'gamma':gamma,
						  'colsample_bytree':colsample_bytree}
            
            
            for params in list(ParameterGrid(param_grid)):
                
                model = XGBClassifier(subsample=params['subsample'],
                                      learning_rate=params['learning_rate'],
                                      max_depth=params['max_depth'],
                                      min_child_weight=params['min_child_weight'],
                                      n_estimators=params['n_estimators'],
                                      gamma=params['gamma'],
									  colsample_bytree=params['colsample_bytree'],
                                      scale_pos_weight=scale_pos_weight,
                                      random_state=random_state,
                                      verbosity=0)
                
                hyper_parameters = OrderedDict({'algorithm':algorithm_name,
                                                'subsample':params['subsample'],
                                                'learning_rate':params['learning_rate'],
                                                'max_depth':params['max_depth'],
                                                'n_estimators':params['n_estimators'],
                                                'min_child_weight':params['min_child_weight'],
                                                'gamma':params['gamma'],
												'colsample_bytree':params['colsample_bytree'],
                                                'scale_pos_weight':scale_pos_weight,
                                                'random_state':random_state})
    
                # model fitting
                model.fit(x_train, y_train, eval_set=eval_set, early_stopping_rounds=500,verbose=False)
    
                # Calculating various metrics
                model_metrics = cls.__model_metrics(model=model,
                                                          x_train=x_train,x_test=x_test,
                                                          y_train=y_train,y_test=y_test,
                                                          iteration_number=iteration_number,
                                                          test_calculation_method=test_calculation_method,
                                                          include_plots=include_plots,
                                                          save_folder=save_folder)
    
                hyper_parameters.update(model_metrics)
                grid_search_list.append(hyper_parameters)
                model_object_list.append(model)
                
                model_parameters_and_metrics = pd.DataFrame({k: [v] for k, v in hyper_parameters.items()})
                
                
                filename = "Iteration_"+str(iteration_number)
                
                # Variable importance score
                predictors = list(x_train.columns)
                feature_importance = pd.DataFrame({'Variable': predictors,"VariableImportance":model.feature_importances_})
                feature_importance.sort_values(by=['VariableImportance'],ascending=False,inplace=True)
                
                cls.__write_grid_search_iterations_results(filename=filename,
                                                parametersList = list(param_grid.keys()),
                                                model_metrics = model_metrics,
                                                model_parameters_and_metrics = model_parameters_and_metrics,
                                                feature_importance = feature_importance,
                                                save_folder=save_folder)
    
                print('Iteration Number '+str(iteration_number)+' ___________________________________________________________________')
                iteration_number += 1
        
            grid_search_results = pd.DataFrame(grid_search_list)
            
            metric_start_index = list(grid_search_results.columns).index('train_AUC_ROC')
            #no_of_params = len(list(param_grid.keys()))
            new_column_names = [re.sub('\_\.','',column_name) for column_name in grid_search_results.columns]
            new_column_names = [re.sub('\_\.','',column_name) for column_name in new_column_names]
            grid_search_results.columns = new_column_names
            
            filename = algorithm_name + "_GridSearchResults"
            filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
            filename += ".xlsx"
            
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 filename), engine = 'openpyxl')
            
            grid_search_results.to_excel(writer,sheet_name='Gridsearch',index_label='IterationNumber',startrow = 4, startcol =1)
            
            if 'table_header' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_header)
            if 'table_title' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_title)
            if 'table_cells' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_cells)
            if 'cell_highlight1' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
            
            # Formatting: GridSearch--------------------------------------------------------------
            sheet = writer.sheets['Gridsearch']
            img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
            img.anchor = ef.generate_excel_cell_name(row_number = 1,
                                                     column_number = 2)
            sheet.add_image(img)
            sheet.cell(row=4, column=2).value = 'Model Parameters'
            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=metric_start_index+2)
            
            metricList = ['AUC_ROC','Accuracy','Precision','Recall','F1_score',
                          'Gini_index','break','Capture_Decile1','Capture_Decile2','Capture_Decile3',
                          'MaxKS','MaxLift','Lift_Decile1','Lift_Decile2','Lift_Decile3',
                          'Population%_Decile1','Population%_Decile2','Population%_Decile3']
            
            max_column_width = max([len(x) for x in metricList ])
            
            metric_index=0
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                sheet.cell(row=4, column=i).value = metricList[metric_index]
                metric_index+=1
                sheet.cell(row=5, column=i).value = 'Train'
                sheet.cell(row=5, column=i+1).value = 'Test'
                
            # table headers
            rows = [4,5] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]
            
            for i in range(0,len(rows)):
                for column in columns[i]:
                    if rows[i]==4:
                        sheet[column+str(rows[i])].style = 'table_title'
                    else:
                        sheet[column+str(rows[i])].style = 'table_header'
                    sheet.column_dimensions[column].width = max_column_width 
#            
            # table cells
            rows = list(range(6,len(grid_search_results)+6))
            columns = [[ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]*len(rows)
                        
#           
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_cells'
                    sheet[column+str(rows[i])].number_format = '0.000'
                    
#                    if (column in ['C','E','F']) and (rows[i]>=25):
#                        sheet[column+str(rows[i])].number_format = '0.00\%'
#                    elif i == 0:
#                        pass
#                    else:
#                        sheet[column+str(rows[i])].number_format = '0.00'
            
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                sheet.merge_cells(start_row=4, start_column=i, end_row=4, end_column=i+1)
#            
#            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=len(parametersList)+1)
#            sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
#            sheet.merge_cells(start_row=17, start_column=2, end_row=17, end_column=6)
#            sheet.merge_cells(start_row=22, start_column=2, end_row=22, end_column=7)
#            sheet.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
#            sheet.merge_cells(start_row=23, start_column=5, end_row=23, end_column=7)
                    
            
            sheet.sheet_view.showGridLines = False
            writer.save()
            writer.close()
            # -------------------------------------------------------------------------------------------
            
            print("\nGrid search iteration files stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory,save_folder))
            
            print("\nGrid search results summary stored in location: \n"+ os.path.join(os.getcwd(),model_tuning.__save_directory))
            print("\nFile Name : \n" + filename)
            
            
            
            grid_search_results['IterationNumber'] = grid_search_results.index
            return grid_search_results, model_object_list
        except TypeError as te:
            print('ERROR:Incorrect datatype of one or more parameters entered. Please check the documentation for detailed understanding.')
            print(str(te))
        #except ValueError as ve:
        #    print('ERROR:',str(ve))
#    
           
    
    @classmethod
    def grid_search_random_forest(cls,
                                  x_train,x_test,
                                  y_train,y_test,
                                  n_estimators = [100],criterion = ['gini'],
                                  max_depth = [None],min_samples_split = [2],
                                  min_samples_leaf = [1],max_features = ['auto'],
                                  bootstrap = [True],
                                  random_state = 0,
                                  test_calculation_method = "ScoreCutOff",
                                  save_charts = True,
                                  include_plots = True):
        
        """This function performs grid search for hyperparameter tuning for RandomForestClassifier.
        
        Please note that by default this function creates a csv for every iteration and hence may be time consuming.
        Set the save_charts flag to False to disable the save feature. 
        
        Parameters
        ----------
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
            Training data, where n_samples is the number of samples and n_features is the number of features.
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
            Test data, where n_samples is the number of samples and n_features is the number of features.		
		
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
        
        n_estimators: list of integers, default=[100]
            Specifies number of trees in the forest.
        
        criterion: list of strings, default=["gini"]
            Function to measure the quality of a split.
            Accepted values: {'gini', 'entropy'}.
        
        max_depth: list of integers, default=[None]
            The maximum depth of the tree. 
        
        min_samples_split: list of integers , default=[2]
            Minimum number of samples required to split an internal node.
        
        min_samples_leaf: list of integers, default=[1]
            Minimum number of samples required to be at a leaf node.
        
        max_features: list of integers, default=['auto']
            Number of features to consider when looking for the best split. 
            'auto' means it takes sqrt(n_features).
        
        bootstrap: list of boolean, default=[True]
            Whether bootstrap samples are used when building trees. If False, the whole dataset is used to build each tree.
            
        random_state: integer, default=0
            Random number seed.
		
        test_calculation_method: string, default="ScoreCutOff"
            Whether to divide test set into equal size deciles or divide into deciles using score cuts.
            Accepted values: {"ScoreCutOff", "EqualDeciles"}.
        
        save_charts: boolean
            default = True
            True, to save KS charts for every iteration.
        
        include_plots: boolean, default=True
            True, to save KS Lift plot for every iteration.
        
        Returns
        -------
        
        grid_search_results: dataframe
            Contains parameter values and metric scores for every iteration.
        
        model_object_list: list
            Contains the model objects generated for each iteration
         
        Additional files generated  
        --------------------------
        
        IterationFolder: KS charts for every iteration.
        
        RandomForest_GridSearchResults.csv: csv file containing grid_search_results.

        Usage
        -----
        This functions results one summary csv 'GradientBoosting_GridSearchResults' which has all the parameter sets and the corresponding accuracies. One can select the best parameter set
        based on the accuracy results. Once the optimal parameter set is chosen, one can look at all the results and plots related to the same iteration 
        in the 'Iteration folder' which stores results for each iteration in a seperate file. So let's say iteration 10 is the optimal one, iteration 10 file can be 
        used to consume the detailed formatted results related to that iteration. Model object with optimal parameters can also be selected from the 'model_object_list'
        for further analysis. So if 10th iteration is the optimal one, one can select 10th element of this list to get the corresponding model object.
       
       
        
        Example
        -------
        >>> from model_tuning import model_tuning
        >>> grid_search_results,model_object_list = model_tuning.grid_search_random_forest(x_train,x_test,y_train,y_test,max_depth=[10,30,50,70,90],
                                                                                           n_estimators = [200,600])

        """
        
        if (np.isfinite(x_train).values.all()==False) or (np.isfinite(y_train).values.all()==False) or (np.isfinite(x_test).values.all()==False) or (np.isfinite(y_test).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
            
        algorithm_name = 'RandomForest'
        save_folder = algorithm_name+'_'+'Iterations'
        save_folder = ef.create_version(base_filename=save_folder, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        status = ef.create_folder(save_folder,path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        if status==False:
            return
            
        try:
            label_encoder_object = LabelEncoder()
            y_train = label_encoder_object.fit_transform(y_train)
            y_test = label_encoder_object.fit_transform(y_test)
    
            
            iteration_number = 0
            grid_search_list = []
            model_object_list = []
    
    
            print("Running iterations..........\n")
            
            param_grid = {'n_estimators':n_estimators,
                          'criterion':criterion,
                          'max_depth':max_depth,
                          'min_samples_split':min_samples_split,
                          'min_samples_leaf':min_samples_leaf,
                          'max_features':max_features,
                          'bootstrap':bootstrap}
            
            
            for params in list(ParameterGrid(param_grid)):
                
                model = RandomForestClassifier(n_estimators = params['n_estimators'],
                                               criterion = params['criterion'],
                                               max_depth = params['max_depth'],
                                               min_samples_split = params['min_samples_split'],
                                               min_samples_leaf = params['min_samples_leaf'],
                                               max_features = params['max_features'],
                                               bootstrap = params['bootstrap'],
                                               random_state=random_state)
    
                hyper_parameters = OrderedDict({'algorithm':algorithm_name,
                                                'n_estimators':params['n_estimators'],
                                                'criterion':params['criterion'],
                                                'max_depth':params['max_depth'],
                                                'min_samples_split':params['min_samples_split'],
                                                'min_samples_leaf':params['min_samples_leaf'],
                                                'max_features':params['max_features'],
                                                'bootstrap':params['bootstrap'],
                                                'random_state':random_state})
                # model fitting
                model.fit(x_train, y_train) 
    
    
                # Calculating various metrics
                
                model_metrics = cls.__model_metrics(model=model,
                                                    x_train=x_train,x_test=x_test,
                                                    y_train=y_train,y_test=y_test,
                                                    iteration_number=iteration_number,
                                                    test_calculation_method=test_calculation_method,
                                                    include_plots=include_plots,
                                                    save_folder=save_folder)
    
    
    
                hyper_parameters.update(model_metrics)
                grid_search_list.append(hyper_parameters)
                model_object_list.append(model)
    
                model_parameters_and_metrics = pd.DataFrame({k: [v] for k, v in hyper_parameters.items()})
    
                filename = "Iteration_"+str(iteration_number)
                
                # Variable importance score
                predictors = list(x_train.columns)
                feature_importance = pd.DataFrame({'Variable': predictors,"VariableImportance":model.feature_importances_})
                feature_importance.sort_values(by=['VariableImportance'],ascending=False,inplace=True)

                cls.__write_grid_search_iterations_results(filename=filename,
                                                parametersList = list(param_grid.keys()),
                                                model_metrics = model_metrics,
                                                model_parameters_and_metrics = model_parameters_and_metrics,
                                                feature_importance = feature_importance,
                                                save_folder=save_folder)
    
                print('Iteration Number '+str(iteration_number)+' ___________________________________________________________________')
                iteration_number += 1
        
            grid_search_results = pd.DataFrame(grid_search_list)
            
            metric_start_index = list(grid_search_results.columns).index('train_AUC_ROC')
            #no_of_params = len(list(param_grid.keys()))
            new_column_names = [re.sub('\_\.','',column_name) for column_name in grid_search_results.columns]
            new_column_names = [re.sub('\_\.','',column_name) for column_name in new_column_names]
            grid_search_results.columns = new_column_names
            
            filename = algorithm_name + "_GridSearchResults"
            filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
            filename += ".xlsx"
            
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 filename), engine = 'openpyxl')
            
            grid_search_results.to_excel(writer,sheet_name='Gridsearch',index_label='IterationNumber',startrow = 4, startcol =1)
            
            if 'table_header' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_header)
            if 'table_title' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_title)
            if 'table_cells' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_cells)
            if 'cell_highlight1' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
            
            # Formatting: GridSearch--------------------------------------------------------------
            sheet = writer.sheets['Gridsearch']
            img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
            img.anchor = ef.generate_excel_cell_name(row_number = 1,
                                                     column_number = 2)
            sheet.add_image(img)
            sheet.cell(row=4, column=2).value = 'Model Parameters'
            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=metric_start_index+2)
            
            metricList = ['AUC_ROC','Accuracy','Precision','Recall','F1_score',
                          'Gini_index','Capture_Decile1','Capture_Decile2','Capture_Decile3',
                          'MaxKS','MaxLift','Lift_Decile1','Lift_Decile2','Lift_Decile3',
                          'Population%_Decile1','Population%_Decile2','Population%_Decile3']
            
            max_column_width = max([len(x) for x in metricList ])
            
            metric_index=0
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                sheet.cell(row=4, column=i).value = metricList[metric_index]
                metric_index+=1
                sheet.cell(row=5, column=i).value = 'Train'
                sheet.cell(row=5, column=i+1).value = 'Test'
                
            # table headers
            rows = [4,5] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]
            
            for i in range(0,len(rows)):
                for column in columns[i]:
                    if rows[i]==4:
                        sheet[column+str(rows[i])].style = 'table_title'
                    else:
                        sheet[column+str(rows[i])].style = 'table_header'
                    sheet.column_dimensions[column].width = max_column_width 
#            
            # table cells
            rows = list(range(6,len(grid_search_results)+6))
            columns = [[ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]*len(rows)
                        
#           
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_cells'
                    sheet[column+str(rows[i])].number_format = '0.000'
                    
#                    if (column in ['C','E','F']) and (rows[i]>=25):
#                        sheet[column+str(rows[i])].number_format = '0.00\%'
#                    elif i == 0:
#                        pass
#                    else:
#                        sheet[column+str(rows[i])].number_format = '0.00'
            
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                sheet.merge_cells(start_row=4, start_column=i, end_row=4, end_column=i+1)
#            
#            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=len(parametersList)+1)
#            sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
#            sheet.merge_cells(start_row=17, start_column=2, end_row=17, end_column=6)
#            sheet.merge_cells(start_row=22, start_column=2, end_row=22, end_column=7)
#            sheet.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
#            sheet.merge_cells(start_row=23, start_column=5, end_row=23, end_column=7)
                    
            
            sheet.sheet_view.showGridLines = False
            writer.save()
            writer.close()
            # -------------------------------------------------------------------------------------------
            
            print("\nGrid search iteration files stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory,save_folder))
            
            print("\nGrid search results summary stored in location: \n"+ os.path.join(os.getcwd(),model_tuning.__save_directory))
            print("\nFile Name : \n" + filename)
            
            
            
            grid_search_results['IterationNumber'] = grid_search_results.index
            return grid_search_results, model_object_list
        except TypeError as te:
            print('ERROR:Incorrect datatype of one or more parameters entered. Please check the documentation for detailed understanding.')
            print(str(te))
        except ValueError as ve:
            print('ERROR:',str(ve))
#    
    
    
    @classmethod
    def grid_search_GBM(cls,
                        x_train, x_test,
                        y_train, y_test,
                        subsample = [1.0], learning_rate = [0.1],
                        n_estimators = [100], criterion = ['friedman_mse'],
                        max_depth = [3], min_samples_split = [2],
                        min_samples_leaf = [1], max_features = [None],
                        random_state = 0,                        
                        test_calculation_method = "ScoreCutOff",
                        save_charts = True,include_plots = True):
        
        """This function performs grid search for hyperparameter tuning for GradientBoosting.
        
        Please note that by default this function creates a csv for every iteration and hence may be time consuming.
        Set the save_charts flag to False to disable the save feature. 
        
        Parameters
        ----------
        x_train: array-like, sparse matrix of shape (n_samples, n_features)
            Training data, where n_samples is the number of samples and n_features is the number of features.
        
        x_test: array-like, sparse matrix of shape (n_samples, n_features)
            Test data, where n_samples is the number of samples and n_features is the number of features.		
		
        y_train: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of train data for supervised learning problems.
        
        y_test: array-like of shape (n_samples,) or (n_samples, n_outputs)
            The target variable of test data for supervised learning problems.
            
        subsample: list of float values, default=[1.0]
            Specifies fraction of samples to be used for fitting the individual base learners.
        
        learning_rate: list of float values, default=[0.1]
            Specifies the step size at each iteration.
            
        n_estimators: list of integers, default=[100]
            Specifies number of trees in the forest.
            
        criterion: list of strings, default=["gini"]
            Accepted values: {“gini”, “entropy”}.
            Function to measure the quality of a split.
        
        max_depth: list of integers, default=[None]
            The maximum depth of the tree.
        
        min_samples_split: list of integers, default=[2]
            Minimum number of samples required to split an internal node.
        
        min_samples_leaf: list of integers, default=[1]
            Minimum number of samples required to be at a leaf node.
        
        max_features: list of integers, default=[None]
            Number of features to consider when looking for the best split. By default is n_features.
            
        random_state: integer, default=0
            Random number seed.
        
        test_calculation_method: string, default="ScoreCutOff"
            Accepted values: {"ScoreCutOff", "EqualDeciles"}.
            Whether to divide test set into equal size deciles or divide into deciles using score cuts.
            
        save_charts: boolean
            default = True
            True, to save KS charts for every iteration.
        
        include_plots: boolean, default = True
            True, to save KS Lift plot for every iteration.
        
        Returns
        -------
        
        grid_search_results: dataframe
            Contains parameter values and metric scores for every iteration.
            
        model_object_list: list
            Contains the model objects generated for each iteration

        
        Additional files generated  
        --------------------------
        
        Iteration folder: KS charts and plots(optional) for every iteration.
        
        GradientBoosting_GridSearchResults.csv: csv file containing grid_search_results.

        Usage
        -----
        This functions results one summary csv 'GradientBoosting_GridSearchResults' which has all the parameter sets and the corresponding accuracies. One can select the best parameter set
        based on the accuracy results. Once the optimal parameter set is chosen, one can look at all the results and plots related to the same iteration 
        in the 'Iteration folder' which stores results for each iteration in a seperate file. So let's say iteration 10 is the optimal one, iteration 10 file can be 
        used to consume the detailed formatted results related to that iteration. Model object with optimal parameters can also be selected from the 'model_object_list'
        for further analysis. So if 10th iteration is the optimal one, one can select 10th element of this list to get the corresponding model object.
       
       
        
        Example
        -------
        
        >>> from model_tuning import model_tuning
        >>> grid_search_results,model_object_list = model_tuning.grid_search_GBM(x_train,x_test,y_train,y_test,subsample=[1.0],learning_rate=[0.1,0.2,0.3],
                                         n_estimators = [50,100])
        
        """
        
        if (np.isfinite(x_train).values.all()==False) or (np.isfinite(y_train).values.all()==False) or (np.isfinite(x_test).values.all()==False) or (np.isfinite(y_test).values.all()==False):
            raise Exception('There may be missing values or inf values present in your data.\nPlease perform the necessary imputation using the impute function from the EDA module first.')
            
        algorithm_name = 'GradientBoosting'
        save_folder = algorithm_name+'_'+'Iterations'
        save_folder = ef.create_version(base_filename=save_folder, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        status = ef.create_folder(save_folder,path = os.path.join(os.getcwd(),model_tuning.__save_directory))
        if status==False:
            return
        
        try:
        
            label_encoder_object = LabelEncoder()
            y_train = label_encoder_object.fit_transform(y_train)
            y_test = label_encoder_object.fit_transform(y_test)
            
            iteration_number = 0
            grid_search_list = []
            model_object_list = []
    
    
            print("Running iterations..........\n")
            
            param_grid = {'subsample':subsample,
                          'learning_rate':learning_rate,
                          'n_estimators':n_estimators,
                          'criterion':criterion,
                          'max_depth':max_depth,
                          'min_samples_split':min_samples_split,
                          'min_samples_leaf':min_samples_leaf,
                          'max_features':max_features}
            
            
            for params in list(ParameterGrid(param_grid)):
                model = GradientBoostingClassifier(subsample=params['subsample'],
                                                   learning_rate=params['learning_rate'],
                                                   n_estimators=params['n_estimators'],
                                                   criterion=params['criterion'],
                                                   max_depth=params['max_depth'],
                                                   min_samples_split=params['min_samples_split'],
                                                   min_samples_leaf=params['min_samples_leaf'],
                                                   max_features=params['max_features'],
                                                   random_state=random_state)
                
                hyper_parameters = OrderedDict({'algorithm':algorithm_name,
                                                'subsample':params['subsample'],
                                                 'learning_rate':params['learning_rate'],
                                                 'n_estimators':params['n_estimators'],
                                                 'criterion':params['criterion'],
                                                 'max_depth':params['max_depth'],
                                                 'min_samples_split':params['min_samples_split'],
                                                 'min_samples_leaf':params['min_samples_leaf'],
                                                 'max_features':params['max_features'],
                                                 'random_state':random_state})
                # model fitting
                model.fit(x_train, y_train) 
    
    
                # Calculating various metrics
                
                model_metrics = cls.__model_metrics(model=model,
                                                    x_train=x_train,x_test=x_test,
                                                    y_train=y_train,y_test=y_test,
                                                    iteration_number=iteration_number,
                                                    test_calculation_method=test_calculation_method,
                                                    include_plots=include_plots,
                                                    save_folder=save_folder)
    
    
                hyper_parameters.update(model_metrics)
                
                grid_search_list.append(hyper_parameters)
                model_object_list.append(model)
    
                model_parameters_and_metrics = pd.DataFrame({k: [v] for k, v in hyper_parameters.items()})
    
                filename = "Iteration_"+str(iteration_number)
                
                # Variable importance score
                predictors = list(x_train.columns)
                feature_importance = pd.DataFrame({'Variable': predictors,"VariableImportance":model.feature_importances_})
                feature_importance.sort_values(by=['VariableImportance'],ascending=False,inplace=True)
                
                cls.__write_grid_search_iterations_results(filename=filename,
                                                parametersList = list(param_grid.keys()),
                                                model_metrics = model_metrics,
                                                model_parameters_and_metrics = model_parameters_and_metrics,
                                                feature_importance = feature_importance,
                                                save_folder=save_folder)
    
                print('Iteration Number '+str(iteration_number)+' ___________________________________________________________________')
                iteration_number += 1
        
            grid_search_results = pd.DataFrame(grid_search_list)
            
            metric_start_index = list(grid_search_results.columns).index('train_AUC_ROC')
            #no_of_params = len(list(param_grid.keys()))
            new_column_names = [re.sub('\_\.','',column_name) for column_name in grid_search_results.columns]
            new_column_names = [re.sub('\_\.','',column_name) for column_name in new_column_names]
            grid_search_results.columns = new_column_names
            
            filename = algorithm_name + "_GridSearchResults"
            filename = ef.create_version(base_filename=filename, path = os.path.join(os.getcwd(),model_tuning.__save_directory))
            filename += ".xlsx"
            
            writer = pd.ExcelWriter(os.path.join(os.getcwd(),
                                                 model_tuning.__save_directory,
                                                 filename), engine = 'openpyxl')
            
            grid_search_results.to_excel(writer,sheet_name='Gridsearch',index_label='IterationNumber',startrow = 4, startcol =1)
            
            if 'table_header' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_header)
            if 'table_title' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_title)
            if 'table_cells' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_table_cells)
            if 'cell_highlight1' not in writer.book.named_styles:
                writer.book.add_named_style(model_tuning.__format_cell_highlight1)
        
            
            # Formatting: Gridsearch--------------------------------------------------------------
            sheet = writer.sheets['Gridsearch']
            img = Image(os.path.join(os.getcwd(),'src','ks_logo.png'))
            img.anchor = ef.generate_excel_cell_name(row_number = 1,
                                                     column_number = 2)
            sheet.add_image(img)
            sheet.cell(row=4, column=2).value = 'Model Parameters'
            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=metric_start_index+2)
            
            metricList = ['AUC_ROC','Accuracy','Precision','Recall','F1_score',
                          'Gini_index','Break','Capture_Decile1','Capture_Decile2','Capture_Decile3',
                          'MaxKS','MaxLift','Lift_Decile1','Lift_Decile2','Lift_Decile3',
                          'Population%_Decile1','Population%_Decile2','Population%_Decile3']
            
            max_column_width = max([len(x) for x in metricList ])
            
            metric_index=0
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                
                sheet.cell(row=4, column=i).value = metricList[metric_index]
                metric_index+=1
                sheet.cell(row=5, column=i).value = 'Train'
                sheet.cell(row=5, column=i+1).value = 'Test'
                
            # table headers
            rows = [4,5] 
            columns = [ [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)],
                        [ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]
            
            for i in range(0,len(rows)):
                for column in columns[i]:
                    if rows[i]==4:
                        sheet[column+str(rows[i])].style = 'table_title'
                    else:
                        sheet[column+str(rows[i])].style = 'table_header'
                    sheet.column_dimensions[column].width = max_column_width 
#            
            # table cells
            rows = list(range(6,len(grid_search_results)+6))
            columns = [[ef.generate_excel_cell_name(None,column_number) for column_number in range(2,len(grid_search_results.columns)+3)]]*len(rows)
                        
#           
            for i in range(0,len(rows)):
                for column in columns[i]:
                    sheet[column+str(rows[i])].style = 'table_cells'
                    sheet[column+str(rows[i])].number_format = '0.000'
                    
#                    if (column in ['C','E','F']) and (rows[i]>=25):
#                        sheet[column+str(rows[i])].number_format = '0.00\%'
#                    elif i == 0:
#                        pass
#                    else:
#                        sheet[column+str(rows[i])].number_format = '0.00'
            
            for i in range(metric_start_index+3,len(grid_search_results.columns)+2,2):
                sheet.merge_cells(start_row=4, start_column=i, end_row=4, end_column=i+1)
#            
#            sheet.merge_cells(start_row=4, start_column=2, end_row=4, end_column=len(parametersList)+1)
#            sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=5)
#            sheet.merge_cells(start_row=17, start_column=2, end_row=17, end_column=6)
#            sheet.merge_cells(start_row=22, start_column=2, end_row=22, end_column=7)
#            sheet.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
#            sheet.merge_cells(start_row=23, start_column=5, end_row=23, end_column=7)
                    
            
            sheet.sheet_view.showGridLines = False
            writer.save()
            writer.close()
            # -------------------------------------------------------------------------------------------
            
            print("\nGrid search iteration files stored at: \n"+os.path.join(os.getcwd(),model_tuning.__save_directory,save_folder))
            
            print("\nGrid search results summary stored in location: \n"+ os.path.join(os.getcwd(),model_tuning.__save_directory))
            print("\nFile Name : \n" + filename)
            
            
            
            grid_search_results['IterationNumber'] = grid_search_results.index
            return grid_search_results, model_object_list
        except TypeError as te:
            print('ERROR:Incorrect datatype of one or more parameters entered. Please check the documentation for detailed understanding.')
            print(str(te))
        except ValueError as ve:
            print('ERROR:',str(ve))
#         